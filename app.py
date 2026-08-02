# -*- coding: utf-8 -*-
"""
大包集团经营概览 v2.0
=====================
统一游戏币记账 | 库存成本核算 | 预收货池缓冲 | 资产/利润总览

原则:
  - 所有交易以游戏币计量（现金买点卡按比例折算）
  - 总资产 = 游戏币余额 + 库存总成本 + 预收货池余额
  - 净利润 = 收入总和 - 点卡费用
  - 收货不产生费用，只增加库存资产；卖出时确认收入
"""

import os
import shutil
import threading
import time
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook

# ==================== 配置 ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
DATA_FILE = os.path.join(DATA_DIR, "mhxy_data.xlsx")
BACKUP_KEEP_DAYS = 30  # 保留最近 30 天的备份

app = Flask(__name__)


# ==================== Excel 数据库 ====================
class ExcelDB:
    def __init__(self, filepath):
        self.filepath = filepath
        self._ensure_file()

    # 所有表及其列定义（用于自动创建和迁移）
    SHEETS = {
        "config": ["key", "value"],
        "inventory": ["id", "name", "quantity", "avg_cost", "total_cost",
                       "market_price", "last_sell_price", "status", "created_at"],
        "pre_receipt_batches": ["id", "batch_name", "total_amount", "allocated_amount",
                                 "status", "note", "created_at"],
        "pre_receipt_items": ["id", "batch_id", "name", "quantity", "unit_cost",
                               "total_cost", "created_at"],
        "transactions": ["id", "date", "type", "category", "description",
                          "amount", "balance_after", "profit", "ref_id", "created_at"],
        "daily_time": ["id", "date", "login_time", "logout_time", "hours",
                        "points_per_hour", "currency_per_point", "account_count",
                        "total_cost", "note", "created_at"],
        "currency_prices": ["id", "date", "time", "price", "note", "created_at"],
    }

    def _ensure_file(self):
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        if not os.path.exists(self.filepath):
            # 全新创建
            wb = Workbook()
            first = True
            for name, headers in self.SHEETS.items():
                if first:
                    ws = wb.active
                    ws.title = name
                    first = False
                else:
                    ws = wb.create_sheet(name)
                ws.append(headers)
            wb.save(self.filepath)
        else:
            # 自动迁移：补上缺失的 sheet
            wb = load_workbook(self.filepath)
            existing = wb.sheetnames
            for name, headers in self.SHEETS.items():
                if name not in existing:
                    ws = wb.create_sheet(name)
                    ws.append(headers)
            wb.save(self.filepath)
            wb.close()
        # 确保 config 中有默认值
        self._ensure_config_defaults()

    def _ensure_config_defaults(self):
        cfg = self.get_config()
        defaults = {
            "currency_ratio": "100000",
            "initial_balance": "0",
            "points_per_hour": "6",
            "currency_per_point": "14000",
        }
        for k, v in defaults.items():
            if k not in cfg:
                self.set_config(k, v)

    # ---------- 通用 ----------
    def _read_all(self, sheet_name):
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [c.value for c in ws[1]]
        wb.close()
        return [dict(zip(headers, r)) for r in rows if any(v is not None for v in r)]

    def _append_row(self, sheet_name, values):
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        ws.append(values)
        wb.save(self.filepath)
        wb.close()

    def _next_id(self, sheet_name):
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        max_id = 0
        for r in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if r[0] is not None:
                try: max_id = max(max_id, int(r[0]))
                except: pass
        wb.close()
        return max_id + 1

    def _update_cell(self, sheet_name, row_id, col_idx, value):
        """更新指定 ID 行的某个单元格"""
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == row_id:
                row[col_idx].value = value
                break
        wb.save(self.filepath)
        wb.close()

    def _delete_row(self, sheet_name, row_id):
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == row_id:
                ws.delete_rows(row[0].row, 1)
                break
        wb.save(self.filepath)
        wb.close()

    # ---------- 配置 ----------
    def get_config(self):
        rows = self._read_all("config")
        return {r["key"]: r["value"] for r in rows}

    def set_config(self, key, value):
        wb = load_workbook(self.filepath)
        ws = wb["config"]
        found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == key:
                row[1].value = str(value)
                found = True
                break
        if not found:
            ws.append([key, str(value)])
        wb.save(self.filepath)
        wb.close()

    # ---------- 游戏币余额 ----------
    def get_balance(self):
        """从最后一笔交易的 balance_after 获取当前余额"""
        wb = load_workbook(self.filepath)
        ws = wb["transactions"]
        last = None
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[6] is not None:  # balance_after column
                last = float(r[6])
        wb.close()
        if last is not None:
            return last
        # 无交易记录，取初始余额
        cfg = self.get_config()
        return float(cfg.get("initial_balance", 0))

    # ---------- 交易流水 ----------
    def add_transaction(self, t_date, t_type, category, description, amount, profit=0, ref_id=None):
        """
        amount: 收入为正，支出为负（绝对值）
        返回新的余额
        """
        tid = self._next_id("transactions")
        current = self.get_balance()
        new_balance = round(current + amount, 0)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._append_row("transactions", [
            tid, t_date, t_type, category, description,
            amount, new_balance, profit, ref_id, now
        ])
        return tid, new_balance

    def get_transactions(self, start_date=None, end_date=None, category=None):
        rows = self._read_all("transactions")
        result = []
        for r in rows:
            d = str(r.get("date", ""))[:10]
            if start_date and d < start_date: continue
            if end_date and d > end_date: continue
            if category and r.get("category") != category: continue
            r["id"] = int(r["id"] or 0)
            r["amount"] = float(r["amount"] or 0)
            r["balance_after"] = float(r["balance_after"] or 0)
            r["profit"] = float(r["profit"] or 0)
            result.append(r)
        return result

    def _recalc_balances(self):
        """重算所有交易的 running balance_after"""
        wb = load_workbook(self.filepath)
        ws = wb["transactions"]
        running = float(self.get_config().get("initial_balance", 0))
        # 收集所有行 (id, amount) 按 id 排序
        rows_data = []
        for row in ws.iter_rows(min_row=2):
            tid = row[0].value
            if tid is None: continue
            amt = float(row[5].value or 0)
            rows_data.append((row[0].row, tid, amt))
        rows_data.sort(key=lambda x: x[1])  # sort by id

        for (excel_row, tid, amt) in rows_data:
            running = round(running + amt, 0)
            ws.cell(row=excel_row, column=7).value = running  # balance_after
        wb.save(self.filepath)
        wb.close()
        return running

    def delete_transaction(self, tid):
        """删除交易并修复库存和余额。
           对收货: 退回库存。对卖货: 返还库存。对点卡: 删除在线时间记录。"""
        txns = self.get_transactions()
        txn = next((t for t in txns if t["id"] == tid), None)
        if not txn:
            raise ValueError("交易不存在")

        cat = txn["category"]

        # 收货(单件/批量): 不能直接删库存(可能已卖出)，仅做余额回退
        if cat == "purchase_single":
            # 尝试找到关联库存并扣减
            desc = (txn.get("description") or "")
            # 格式: "收货: name xqty @cost"
            pass  # 库存已在卖货时扣减，删除收货只回退余额

        if cat == "sale":
            # 卖货删除: 返还库存
            desc = (txn.get("description") or "")
            # 格式: "卖出: name xqty @price (成本@cost)"
            import re
            m = re.search(r"卖出: (.+?) x(\d+)", desc)
            if m:
                name = m.group(1)
                qty = int(m.group(2))
                cost_match = re.search(r"成本@([\d,]+)", desc)
                cost = float(cost_match.group(1).replace(",", "")) if cost_match else 0
                self.add_inventory(name, qty, cost)

        if cat == "point_card":
            # 同步删除 daily_time 记录
            wb = load_workbook(self.filepath)
            if "daily_time" in wb.sheetnames:
                ws = wb["daily_time"]
                for row in ws.iter_rows(min_row=2):
                    if row[0].value and int(row[0].value) == (txn.get("ref_id") or 0):
                        ws.delete_rows(row[0].row, 1)
                        break
                wb.save(self.filepath)
            wb.close()

        # 删除交易行
        self._delete_row("transactions", tid)
        # 重算余额
        self._recalc_balances()

    def update_transaction(self, tid, data):
        """更新交易描述、日期（不动金额和库存）"""
        wb = load_workbook(self.filepath)
        ws = wb["transactions"]
        col_map = {"date": 1, "type": 2, "category": 3, "description": 4}
        for row in ws.iter_rows(min_row=2):
            if row[0].value == tid:
                for key, col_idx in col_map.items():
                    if key in data:
                        row[col_idx].value = data[key]
                break
        wb.save(self.filepath)
        wb.close()

    def get_transaction(self, tid):
        """获取单条交易"""
        txns = self.get_transactions()
        return next((t for t in txns if t["id"] == tid), None)

    # ---------- 库存 ----------
    def get_inventory(self, status="active"):
        items = self._read_all("inventory")
        result = []
        for i in items:
            if status and i.get("status") != status: continue
            i["id"] = int(i["id"] or 0)
            i["quantity"] = int(i["quantity"] or 0)
            i["avg_cost"] = float(i["avg_cost"] or 0)
            i["total_cost"] = float(i["total_cost"] or 0)
            i["market_price"] = float(i["market_price"] or 0)
            i["last_sell_price"] = float(i["last_sell_price"] or 0)
            result.append(i)
        return result

    def add_inventory(self, name, quantity, unit_cost):
        """添加库存（单件收货或批量拆分后入库）。
           使用移动平均成本法：合并同名物品。"""
        total_cost = quantity * unit_cost
        # 查找同名同状态物品
        existing = [i for i in self.get_inventory("active") if i["name"] == name]
        if existing:
            item = existing[0]
            new_qty = item["quantity"] + quantity
            new_total = item["total_cost"] + total_cost
            new_avg = round(new_total / new_qty, 0)
            self._update_cell("inventory", item["id"], 2, new_qty)      # quantity
            self._update_cell("inventory", item["id"], 3, new_avg)      # avg_cost
            self._update_cell("inventory", item["id"], 4, new_total)    # total_cost
            return item["id"]
        else:
            iid = self._next_id("inventory")
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._append_row("inventory", [
                iid, name, quantity, unit_cost, total_cost,
                0, 0, "active", now
            ])
            return iid

    def sell_inventory(self, inv_id, quantity, sell_price):
        """卖出库存：减库存，返回利润"""
        items = self.get_inventory("active")
        item = next((i for i in items if i["id"] == inv_id), None)
        if not item or item["quantity"] < quantity:
            raise ValueError("库存不足")
        cost = item["avg_cost"] * quantity
        revenue = sell_price * quantity
        profit = revenue - cost

        new_qty = item["quantity"] - quantity
        if new_qty <= 0:
            self._update_cell("inventory", inv_id, 2, 0)
            self._update_cell("inventory", inv_id, 4, 0)
            self._update_cell("inventory", inv_id, 7, "sold")  # status
        else:
            new_total = item["avg_cost"] * new_qty
            self._update_cell("inventory", inv_id, 2, new_qty)
            self._update_cell("inventory", inv_id, 4, new_total)
        # 更新最后售价
        self._update_cell("inventory", inv_id, 6, sell_price)
        return profit, cost, item["name"], item["avg_cost"]

    def update_inventory_market(self, inv_id, market_price):
        self._update_cell("inventory", inv_id, 5, market_price)

    # ---------- 预收货池 ----------
    def get_batches(self, status=None):
        batches = self._read_all("pre_receipt_batches")
        result = []
        for b in batches:
            if status and b.get("status") != status: continue
            b["id"] = int(b["id"] or 0)
            b["total_amount"] = float(b["total_amount"] or 0)
            b["allocated_amount"] = float(b["allocated_amount"] or 0)
            result.append(b)
        return result

    def create_batch(self, name, total_amount, note=""):
        bid = self._next_id("pre_receipt_batches")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._append_row("pre_receipt_batches", [
            bid, name, total_amount, 0, "open", note, now
        ])
        return bid

    def get_batch_items(self, batch_id):
        items = self._read_all("pre_receipt_items")
        result = []
        for i in items:
            if int(i.get("batch_id", 0)) != batch_id: continue
            i["id"] = int(i["id"] or 0)
            i["quantity"] = int(i["quantity"] or 0)
            i["unit_cost"] = float(i["unit_cost"] or 0)
            i["total_cost"] = float(i["total_cost"] or 0)
            result.append(i)
        return result

    def add_batch_item(self, batch_id, name, quantity, unit_cost):
        """向预收货批次添加物品"""
        batches = self.get_batches()
        batch = next((b for b in batches if b["id"] == batch_id), None)
        if not batch: raise ValueError("批次不存在")
        if batch["status"] != "open": raise ValueError("批次已关闭")

        total_cost = quantity * unit_cost
        new_allocated = batch["allocated_amount"] + total_cost
        if new_allocated > batch["total_amount"] + 1:  # 1 为浮点容差
            raise ValueError(f"超出批次总额！已分配 {batch['allocated_amount']:,.0f}，"
                           f"本次 {total_cost:,.0f}，总额 {batch['total_amount']:,.0f}")

        iid = self._next_id("pre_receipt_items")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._append_row("pre_receipt_items", [
            iid, batch_id, name, quantity, unit_cost, total_cost, now
        ])
        self._update_cell("pre_receipt_batches", batch_id, 3, new_allocated)
        return iid

    def complete_batch(self, batch_id, absorb_remainder=True):
        """完成预收货批次：剩余金额处理，物品转入库存"""
        batches = self.get_batches()
        batch = next((b for b in batches if b["id"] == batch_id), None)
        if not batch: raise ValueError("批次不存在")

        items = self.get_batch_items(batch_id)
        remainder = batch["total_amount"] - batch["allocated_amount"]

        if remainder > 0 and absorb_remainder:
            # 剩余金额均摊到已有物品
            total_qty = sum(i["quantity"] for i in items)
            if total_qty > 0:
                extra_per = round(remainder / total_qty, 0)
                for item in items:
                    new_unit = item["unit_cost"] + extra_per
                    new_total = item["quantity"] * new_unit
                    self._update_cell("pre_receipt_items", item["id"], 4, new_unit)
                    self._update_cell("pre_receipt_items", item["id"], 5, new_total)
                remainder = 0

        if remainder > 0:
            # 仍有剩余 → 创建"杂项损耗"物品
            self.add_batch_item(batch_id, f"【{batch['batch_name']}】剩余损耗", 1, remainder)

        # 将所有物品转入库存
        final_items = self.get_batch_items(batch_id)
        for item in final_items:
            self.add_inventory(item["name"], item["quantity"], item["unit_cost"])

        # 标记批次完成
        self._update_cell("pre_receipt_batches", batch_id, 4, "done")
        return final_items

    def delete_batch(self, batch_id):
        """删除未完成的批次（退款）"""
        batches = self.get_batches()
        batch = next((b for b in batches if b["id"] == batch_id), None)
        if not batch: return
        # 删除批次下的物品
        for item in self.get_batch_items(batch_id):
            self._delete_row("pre_receipt_items", item["id"])
        self._delete_row("pre_receipt_batches", batch_id)

    def delete_batch_item(self, batch_id, item_id):
        """删除批次中的单个物品，退回已分配金额"""
        batches = self.get_batches()
        batch = next((b for b in batches if b["id"] == batch_id), None)
        if not batch or batch["status"] != "open":
            raise ValueError("批次不存在或已关闭")
        items = self.get_batch_items(batch_id)
        item = next((i for i in items if i["id"] == item_id), None)
        if not item:
            raise ValueError("物品不存在")
        # 退回已分配金额
        new_allocated = batch["allocated_amount"] - item["total_cost"]
        self._update_cell("pre_receipt_batches", batch_id, 3, max(0, new_allocated))
        self._delete_row("pre_receipt_items", item_id)

    def reopen_batch(self, batch_id):
        """重新打开已完成的批次（物品保留在库存中，不回退）"""
        batches = self.get_batches()
        batch = next((b for b in batches if b["id"] == batch_id), None)
        if not batch or batch["status"] != "done":
            raise ValueError("只能重开已完成的批次")
        self._update_cell("pre_receipt_batches", batch_id, 4, "open")

    # ---------- 每日在线时间 ----------
    def get_daily_times(self, start_date=None, end_date=None):
        rows = self._read_all("daily_time")
        result = []
        for r in rows:
            d = str(r.get("date", ""))[:10]
            if start_date and d < start_date: continue
            if end_date and d > end_date: continue
            r["id"] = int(r["id"] or 0)
            r["hours"] = float(r["hours"] or 0)
            r["total_cost"] = float(r["total_cost"] or 0)
            result.append(r)
        return result

    def add_daily_time(self, start_str, end_str, account_count=1, note=""):
        """记录在线时间，自动计算点卡成本。
           start_str/end_str: 'YYYY-MM-DDTHH:MM' 格式 (datetime-local) """
        tid = self._next_id("daily_time")
        # 解析并计算小时
        fmt = "%Y-%m-%dT%H:%M"
        start_dt = datetime.strptime(start_str, fmt)
        end_dt = datetime.strptime(end_str, fmt)
        if end_dt <= start_dt:
            raise ValueError("下线时间必须晚于上线时间")
        hours = round((end_dt - start_dt).total_seconds() / 3600, 2)
        t_date = start_dt.strftime("%Y-%m-%d")
        login_time = start_dt.strftime("%H:%M")
        logout_time = end_dt.strftime("%H:%M")

        cfg = self.get_config()
        pph = float(cfg.get("points_per_hour", 6))
        cpp = float(cfg.get("currency_per_point", 14000))
        total_cost = round(hours * pph * cpp * account_count, 0)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._append_row("daily_time", [
            tid, t_date, login_time, logout_time, hours,
            pph, cpp, account_count, total_cost, note, now
        ])

        # 生成点卡费用交易
        ac_info = f" ({account_count}开)" if account_count > 1 else ""
        date_range = f"{start_dt.strftime('%m-%d %H:%M')} → {end_dt.strftime('%m-%d %H:%M')}"
        self.add_transaction(
            t_date, "expense", "point_card",
            f"点卡: {date_range} ({hours}h × {pph}点/h × {cpp:,.0f}/点{ac_info})",
            -total_cost
        )
        return tid, hours, total_cost

    # ---------- 币价记录 ----------
    def add_currency_price(self, t_date, t_time, price, note=""):
        cid = self._next_id("currency_prices")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._append_row("currency_prices", [cid, t_date, t_time, price, note, now])
        return cid

    def get_currency_prices(self):
        rows = self._read_all("currency_prices")
        result = []
        for r in rows:
            r["id"] = int(r["id"] or 0)
            r["price"] = float(r["price"] or 0)
            result.append(r)
        result.sort(key=lambda x: (str(x.get("date", "")), str(x.get("time", ""))))
        return result

    def get_currency_trend(self):
        """返回每日均价，用于折线图"""
        rows = self.get_currency_prices()
        daily = {}  # date -> [prices]
        for r in rows:
            d = str(r.get("date", ""))[:10]
            if d:
                daily.setdefault(d, []).append(r["price"])
        result = []
        for d in sorted(daily.keys()):
            prices = daily[d]
            avg = round(sum(prices) / len(prices), 0)
            result.append({
                "date": d,
                "avg_price": avg,
                "min_price": min(prices),
                "max_price": max(prices),
                "price_per_wan": round(avg / 3000, 4),  # 每万两单价(元)
                "count": len(prices),
            })
        return result

    def get_latest_price_per_wan(self):
        """获取最近一天的每万两单价，用于实时获利计算"""
        trend = self.get_currency_trend()
        if trend:
            return trend[-1]["price_per_wan"]
        return 0

    # ---------- 汇总 ----------
    def get_summary(self, month=None):
        """资产总览 + 利润表"""
        balance = self.get_balance()
        inventory_items = self.get_inventory("active")
        inventory_cost = sum(i["total_cost"] for i in inventory_items)
        open_batches = self.get_batches("open")
        pool_balance = sum(b["total_amount"] - b["allocated_amount"] for b in open_batches)
        total_assets = balance + inventory_cost + pool_balance

        # 利润表
        if not month:
            month = date.today().strftime("%Y-%m")
        month_start = month + "-01"
        y, m = int(month[:4]), int(month[5:7])
        if m == 12:
            month_end = date(y+1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(y, m+1, 1) - timedelta(days=1)
        month_end_str = month_end.isoformat()

        txns = self.get_transactions(month_start, month_end_str)

        # 收入分类汇总
        def sum_cat(cat):
            return round(sum(t["amount"] for t in txns
                           if t["category"] == cat and t["amount"] > 0), 0)
        def sum_cat_expense(cat):
            return round(sum(abs(t["amount"]) for t in txns
                           if t["category"] == cat and t["amount"] < 0), 0)

        sale_profit = round(sum(t["profit"] for t in txns
                               if t["category"] == "sale"), 0)
        vitality_income = sum_cat("vitality")
        school_income = sum_cat("school")
        random_income = sum_cat("random")
        point_card_expense = sum_cat_expense("point_card")

        total_income = sale_profit + vitality_income + school_income + random_income
        net_profit = total_income - point_card_expense

        # 实时获利(RMB) = 净利润(万) × 每万两单价(元/万)
        price_per_wan = db.get_latest_price_per_wan()
        rmb_profit = round(net_profit / 10000 * price_per_wan, 2) if price_per_wan > 0 else 0

        # 每日趋势
        daily = []
        d = date(y, m, 1)
        today = date.today()
        end = min(month_end, today)
        while d <= end:
            ds = d.isoformat()
            d_txns = [t for t in txns if str(t.get("date", ""))[:10] == ds]
            d_income = sum(t["profit"] for t in d_txns if t["category"] == "sale")
            d_income += sum(t["amount"] for t in d_txns
                          if t["category"] in ("vitality","school","random") and t["amount"] > 0)
            d_expense = sum(abs(t["amount"]) for t in d_txns
                          if t["category"] == "point_card" and t["amount"] < 0)
            daily.append({
                "date": ds,
                "income": round(d_income, 0),
                "expense": round(d_expense, 0),
                "profit": round(d_income - d_expense, 0),
            })
            d += timedelta(days=1)

        # 可用月份
        all_txns = self.get_transactions()
        months_set = sorted(set(str(t.get("date", ""))[:7] for t in all_txns if t.get("date")), reverse=True)
        if month not in months_set:
            months_set.append(month)
            months_set.sort(reverse=True)

        return {
            "assets": {
                "balance": balance,
                "inventory_cost": inventory_cost,
                "pool_balance": pool_balance,
                "total": total_assets,
                "open_batches": len(open_batches),
            },
            "profit": {
                "month": month,
                "sale_profit": sale_profit,
                "vitality_income": vitality_income,
                "school_income": school_income,
                "random_income": random_income,
                "total_income": total_income,
                "point_card_expense": point_card_expense,
                "net_profit": net_profit,
                "price_per_wan": price_per_wan,
                "rmb_profit": rmb_profit,
            },
            "daily_trend": daily,
            "available_months": months_set,
        }


# ==================== 全局实例 ====================
db = ExcelDB(DATA_FILE)


# ==================== 自动备份 ====================
_last_backup_date = None  # 记录上次备份的日期，防止重复

def backup_data():
    """备份数据到 backups/ 目录，保留最近 N 天"""
    global _last_backup_date
    today_str = date.today().isoformat()
    if _last_backup_date == today_str:
        return  # 今天已经备份过了
    if not os.path.exists(DATA_FILE):
        return  # 还没数据，不备份

    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup_name = f"mhxy_data_{today_str}.xlsx"
    backup_path = os.path.join(BACKUP_DIR, backup_name)

    try:
        shutil.copy2(DATA_FILE, backup_path)
        _last_backup_date = today_str
        print(f"[备份] {datetime.now().strftime('%H:%M:%S')} → {backup_name}")

        # 清理超过 30 天的旧备份
        cutoff = date.today() - timedelta(days=BACKUP_KEEP_DAYS)
        for fname in os.listdir(BACKUP_DIR):
            if fname.startswith("mhxy_data_") and fname.endswith(".xlsx"):
                try:
                    fdate_str = fname.replace("mhxy_data_", "").replace(".xlsx", "")
                    fdate = date.fromisoformat(fdate_str)
                    if fdate < cutoff:
                        os.remove(os.path.join(BACKUP_DIR, fname))
                        print(f"[备份] 清理旧备份: {fname}")
                except ValueError:
                    pass
    except Exception as e:
        print(f"[备份] 失败: {e}")

def backup_scheduler():
    """后台线程：每分钟检查一次是否需要备份（0:00-0:01 触发）"""
    while True:
        now = datetime.now()
        if now.hour == 0 and now.minute <= 1:
            backup_data()
        time.sleep(60)  # 每分钟检查一次


# ==================== 辅助 ====================
def _int(v, default=0):
    try: return int(v)
    except: return default

def _float(v, default=0):
    try: return float(v)
    except: return default


# ==================== API ====================

@app.route("/")
def index():
    return render_template("index.html")

# ---------- 总览 ----------
@app.route("/api/summary")
def api_summary():
    month = request.args.get("month", "").strip() or None
    return jsonify(db.get_summary(month))

# ---------- 流水 ----------
@app.route("/api/transactions")
def api_transactions():
    return jsonify(db.get_transactions(
        request.args.get("start"),
        request.args.get("end"),
        request.args.get("category"),
    ))

@app.route("/api/transactions/<int:tid>", methods=["DELETE"])
def api_delete_transaction(tid):
    try:
        db.delete_transaction(tid)
        return jsonify({"ok": True, "balance": db.get_balance()})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/transactions/<int:tid>", methods=["PUT"])
def api_update_transaction(tid):
    d = request.get_json()
    db.update_transaction(tid, d)
    # 如果日期变了，重算余额
    if "amount" not in d:
        pass  # 仅改描述/日期，余额不变
    return jsonify({"ok": True})

# ---------- 库存 ----------
@app.route("/api/inventory")
def api_inventory():
    return jsonify(db.get_inventory(request.args.get("status") or "active"))

@app.route("/api/inventory/market", methods=["PUT"])
def api_inventory_market():
    d = request.get_json()
    db.update_inventory_market(_int(d["id"]), _float(d["market_price"]))
    return jsonify({"ok": True})

# ---------- 单件收货 ----------
@app.route("/api/purchase/single", methods=["POST"])
def api_purchase_single():
    d = request.get_json()
    name = d.get("name", "").strip()
    qty = _int(d.get("quantity"), 1)
    unit_cost = _float(d.get("unit_cost"))
    if not name or qty <= 0 or unit_cost <= 0:
        return jsonify({"error": "请填写物品名、数量和单价"}), 400

    total = qty * unit_cost
    # 扣余额
    tid, bal = db.add_transaction(
        d.get("date", date.today().isoformat()),
        "expense", "purchase_single",
        f"收货: {name} x{qty} @{unit_cost:,.0f}",
        -total
    )
    # 加库存
    db.add_inventory(name, qty, unit_cost)
    return jsonify({"ok": True, "balance": bal, "transaction_id": tid})

# ---------- 批量收货 ----------
@app.route("/api/purchase/bulk", methods=["POST"])
def api_purchase_bulk():
    d = request.get_json()
    name = d.get("batch_name", "").strip()
    total = _float(d.get("total_amount"))
    note = d.get("note", "")
    if not name or total <= 0:
        return jsonify({"error": "请填写批次名和总额"}), 400

    # 扣余额
    tid, bal = db.add_transaction(
        d.get("date", date.today().isoformat()),
        "expense", "purchase_bulk",
        f"批量收货: {name} ({total:,.0f})",
        -total
    )
    # 创建预收货批次
    bid = db.create_batch(name, total, note)
    return jsonify({"ok": True, "balance": bal, "batch_id": bid, "transaction_id": tid})

# ---------- 预收货池 ----------
@app.route("/api/pre-receipt")
def api_pre_receipt():
    status = request.args.get("status")
    return jsonify(db.get_batches(status))

@app.route("/api/pre-receipt/<int:bid>/items")
def api_batch_items(bid):
    return jsonify(db.get_batch_items(bid))

@app.route("/api/pre-receipt/<int:bid>/items", methods=["POST"])
def api_add_batch_item(bid):
    d = request.get_json()
    try:
        iid = db.add_batch_item(bid, d["name"], _int(d.get("quantity"), 1),
                                _float(d.get("unit_cost")))
        return jsonify({"ok": True, "id": iid})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/pre-receipt/<int:bid>/complete", methods=["POST"])
def api_complete_batch(bid):
    try:
        items = db.complete_batch(bid)
        return jsonify({"ok": True, "items_count": len(items)})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/pre-receipt/<int:bid>", methods=["DELETE"])
def api_delete_batch(bid):
    # 退款：返还余额
    batches = db.get_batches()
    batch = next((b for b in batches if b["id"] == bid), None)
    if batch and batch["status"] == "open":
        refund = batch["total_amount"] - batch["allocated_amount"]
        if refund > 0:
            db.add_transaction(
                date.today().isoformat(), "income", "random",
                f"退款: 取消批次「{batch['batch_name']}」",
                refund
            )
    db.delete_batch(bid)
    return jsonify({"ok": True})

@app.route("/api/pre-receipt/<int:bid>/items/<int:iid>", methods=["DELETE"])
def api_delete_batch_item(bid, iid):
    try:
        db.delete_batch_item(bid, iid)
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

# ---------- 卖出 ----------
@app.route("/api/sell", methods=["POST"])
def api_sell():
    d = request.get_json()
    inv_id = _int(d.get("inventory_id"))
    qty = _int(d.get("quantity"), 1)
    sell_price = _float(d.get("sell_price"))
    if not inv_id or qty <= 0 or sell_price <= 0:
        return jsonify({"error": "请选择库存物品，填写数量和售价"}), 400

    try:
        profit, cost, name, avg_cost = db.sell_inventory(inv_id, qty, sell_price)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    revenue = sell_price * qty
    tid, bal = db.add_transaction(
        d.get("date", date.today().isoformat()),
        "income", "sale",
        f"卖出: {name} x{qty} @{sell_price:,.0f} (成本@{avg_cost:,.0f})",
        revenue,
        profit=profit
    )
    return jsonify({
        "ok": True, "balance": bal, "profit": profit,
        "revenue": revenue, "cost": cost, "transaction_id": tid
    })

# ---------- 体活 / 师门 / 随机收入 ----------
@app.route("/api/income", methods=["POST"])
def api_income():
    d = request.get_json()
    cat = d.get("category", "")
    if cat not in ("vitality", "school", "random"):
        return jsonify({"error": "无效的收入类型"}), 400

    amount = _float(d.get("amount"))
    desc = d.get("description", "").strip()
    if amount <= 0:
        return jsonify({"error": "金额必须大于0"}), 400

    labels = {"vitality": "体活收入", "school": "师门收入", "random": "随机物品收入"}
    tid, bal = db.add_transaction(
        d.get("date", date.today().isoformat()),
        "income", cat,
        desc or labels[cat],
        amount
    )
    return jsonify({"ok": True, "balance": bal, "transaction_id": tid})

# ---------- 点卡（按在线时间自动计算） ----------
@app.route("/api/pointcard", methods=["POST"])
def api_pointcard():
    d = request.get_json()
    start_str = d.get("start", "").strip()
    end_str = d.get("end", "").strip()
    note = d.get("note", "").strip()
    account_count = _int(d.get("account_count"), 1)

    if not start_str or not end_str:
        return jsonify({"error": "请填写上线和下线时间"}), 400
    if account_count < 1:
        account_count = 1

    try:
        tid, hours, total_cost = db.add_daily_time(start_str, end_str, account_count, note)
    except Exception as e:
        return jsonify({"error": f"时间格式错误: {e}"}), 400

    bal = db.get_balance()
    return jsonify({
        "ok": True, "balance": bal, "hours": hours, "total_cost": total_cost,
        "transaction_id": tid
    })

@app.route("/api/daily-time")
def api_daily_time():
    return jsonify(db.get_daily_times(
        request.args.get("start"),
        request.args.get("end"),
    ))

# ---------- 币价趋势 ----------
@app.route("/api/currency-price")
def api_currency_prices():
    return jsonify(db.get_currency_prices())

@app.route("/api/currency-price/trend")
def api_currency_trend():
    return jsonify(db.get_currency_trend())

@app.route("/api/currency-price", methods=["POST"])
def api_add_currency_price():
    d = request.get_json()
    t_date = d.get("date", date.today().isoformat())
    t_time = d.get("time", datetime.now().strftime("%H:%M"))
    price = _float(d.get("price"))
    note = d.get("note", "").strip()
    if price <= 0:
        return jsonify({"error": "请输入有效价格"}), 400
    cid = db.add_currency_price(t_date, t_time, price, note)
    return jsonify({"ok": True, "id": cid})

# ---------- 初始化余额 ----------
@app.route("/api/balance/init", methods=["POST"])
def api_init_balance():
    d = request.get_json()
    amount = _float(d.get("amount"))
    if amount < 0:
        return jsonify({"error": "余额不能为负"}), 400

    current = db.get_balance()
    if current == 0:
        # 首次初始化：写入一条 amount=0 的流水，balance_after 直接设为初始余额
        db.set_config("initial_balance", str(amount))
        tid = db._next_id("transactions")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        db._append_row("transactions", [
            tid, date.today().isoformat(), "income", "balance_init",
            "初始化游戏币余额: " + str(int(amount)),
            0, amount, 0, None, now
        ])
        return jsonify({"ok": True, "balance": amount})

    # 校正模式：差额调整
    diff = amount - current
    if diff != 0:
        db.add_transaction(
            date.today().isoformat(), "income" if diff > 0 else "expense",
            "balance_init",
            f"余额校正 (原{current:,.0f} → {amount:,.0f})",
            diff
        )
    return jsonify({"ok": True, "balance": amount})

# ---------- 配置 ----------
@app.route("/api/config")
def api_config():
    return jsonify(db.get_config())

@app.route("/api/config", methods=["PUT"])
def api_update_config():
    d = request.get_json()
    for k, v in d.items():
        db.set_config(k, v)
    return jsonify({"ok": True})


# ==================== 启动 ====================
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    print("大包集团经营概览 v2.0 启动中...")
    print(f"数据文件: {DATA_FILE}")
    print(f"备份目录: {BACKUP_DIR}")
    print(f"访问地址: http://localhost:5000")

    # 启动自动备份线程（守护线程，主进程退出时自动结束）
    t = threading.Thread(target=backup_scheduler, daemon=True)
    t.start()

    app.run(debug=True, host="127.0.0.1", port=5000)
